from __future__ import annotations

import csv
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from mailscan2026.core import priority, session_store


EXPORT_COLUMNS = [
    "Exported At",
    "Priority",
    "Status",
    "Category",
    "Sender",
    "Type",
    "Amount",
    "Due Date",
    "Payable?",
    "Needs Review",
    "Confidence",
    "Source PDF",
    "Notes",
    "Review Flags",
]


@dataclass(frozen=True)
class FinanceExportSummary:
    total_rows: int
    exported_rows: int
    payable_rows: int
    payable_total: float
    csv_path: Path
    xlsx_path: Path


def export_finance_files(rows: list[dict[str, str]], reviewed_only: bool = False) -> FinanceExportSummary:
    export_rows = prepare_export_rows(rows, reviewed_only=reviewed_only)
    export_dir = session_store.app_data_dir() / "exports"
    export_dir.mkdir(parents=True, exist_ok=True)

    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    suffix = "reviewed" if reviewed_only else "all"
    csv_path = export_dir / f"finance_export_{suffix}_{stamp}.csv"
    xlsx_path = export_dir / f"finance_export_{suffix}_{stamp}.xlsx"

    write_csv(export_rows, csv_path)
    write_xlsx(export_rows, xlsx_path)

    payable_total = 0.0
    payable_rows = 0
    for row in export_rows:
        if row.get("Payable?") == "Yes":
            amount = priority.parse_amount(row.get("Amount", ""))
            if amount is not None:
                payable_rows += 1
                payable_total += amount

    return FinanceExportSummary(
        total_rows=len(rows),
        exported_rows=len(export_rows),
        payable_rows=payable_rows,
        payable_total=payable_total,
        csv_path=csv_path,
        xlsx_path=xlsx_path,
    )


def prepare_export_rows(rows: list[dict[str, str]], reviewed_only: bool = False) -> list[dict[str, str]]:
    prepared: list[dict[str, str]] = []
    now = datetime.now().isoformat(timespec="seconds")
    for row in rows:
        status = str(row.get("Status", "")).strip()
        if reviewed_only and status.lower() != "reviewed":
            continue
        doc_type = str(row.get("Type", ""))
        amount = str(row.get("Amount", "")).strip()
        payable = "Yes" if "bill / payable" in doc_type.lower() and priority.parse_amount(amount) is not None else "No"
        computed_priority = str(row.get("Priority", "")).strip() or priority.compute_priority(row).label
        prepared.append({
            "Exported At": now,
            "Priority": computed_priority,
            "Status": status,
            "Category": str(row.get("Category", "")),
            "Sender": str(row.get("Sender", "")),
            "Type": doc_type,
            "Amount": amount,
            "Due Date": str(row.get("Due Date", "")),
            "Payable?": payable,
            "Needs Review": str(row.get("Needs Review", "")),
            "Confidence": str(row.get("Confidence", "")),
            "Source PDF": str(row.get("Source PDF", "")),
            "Notes": str(row.get("Notes", "")),
            "Review Flags": str(row.get("Review Flags", "")),
        })
    return prepared


def write_csv(rows: list[dict[str, str]], path: Path) -> None:
    with path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=EXPORT_COLUMNS)
        writer.writeheader()
        writer.writerows(rows)


def write_xlsx(rows: list[dict[str, str]], path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Finance Export"
    ws.append(EXPORT_COLUMNS)

    header_fill = PatternFill("solid", fgColor="D9D9D9")
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill

    for row in rows:
        ws.append([row.get(col, "") for col in EXPORT_COLUMNS])

    priority_fills = {
        "Urgent": PatternFill("solid", fgColor="FFD6D6"),
        "Due Soon": PatternFill("solid", fgColor="FFE4C4"),
        "Review": PatternFill("solid", fgColor="FFF7BF"),
        "Info": PatternFill("solid", fgColor="DCEBFF"),
        "Reviewed": PatternFill("solid", fgColor="D8F5D0"),
        "Ignored": PatternFill("solid", fgColor="EEEEEE"),
    }
    priority_col = EXPORT_COLUMNS.index("Priority") + 1
    for row_index in range(2, ws.max_row + 1):
        label = ws.cell(row=row_index, column=priority_col).value
        fill = priority_fills.get(str(label), None)
        if fill:
            for col_index in range(1, ws.max_column + 1):
                ws.cell(row=row_index, column=col_index).fill = fill

    for col_index, col_name in enumerate(EXPORT_COLUMNS, start=1):
        width = min(max(len(col_name) + 2, 12), 42)
        for row_index in range(2, ws.max_row + 1):
            value = ws.cell(row=row_index, column=col_index).value
            if value:
                width = min(max(width, len(str(value)) + 2), 42)
        ws.column_dimensions[get_column_letter(col_index)].width = width

    ws.freeze_panes = "A2"
    wb.save(path)
